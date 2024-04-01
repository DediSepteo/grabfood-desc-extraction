import tkinter as tk
import html
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl

def find_desc(dish_name):
    itemNameXPATH = f"//p[contains(text(), '{dish_name}')]/../../p"
    try:
        itemName = driver.find_element(By.XPATH, itemNameXPATH)
        ActionChains(driver).move_to_element(itemName).perform()
        desc = itemName.get_attribute("innerHTML")
        decoded_desc = html.unescape(desc)  # Decode HTML entities
        return decoded_desc
    except:
        return " "

def load_url():
    url = url_entry.get()
    driver.get(url)
    time.sleep(5)  # Add some delay to ensure the page loads completely

    # Iterate over dish names from the Excel and find descriptions
    for index, row in excel_data_df.iterrows():
        dish_name = row['Dish Name']
        description = find_desc(dish_name)
        excel_data_df.at[index, 'Description'] = description

    # Load existing Excel file
    wb = openpyxl.load_workbook('Prototype - Copy - Copy.xlsx')

    # Get the sheet with the desired name or create it if it doesn't exist
    if 'Menu Items' in wb.sheetnames:
        sheet = wb['Menu Items']
    else:
        sheet = wb.create_sheet('Menu Items')

    # Write DataFrame to the existing Excel file without overwriting other sheets
    for row_index, row in excel_data_df.iterrows():
        sheet.cell(row=row_index + 2, column=3).value = row['Dish Name']
        sheet.cell(row=row_index + 2, column=4).value = row['Description']

    # Save the modified workbook
    wb.save('Prototype - Copy - Copy.xlsx')
    print("Descriptions updated and saved to 'Prototype - Copy - Copy.xlsx'")

# Initialize Tkinter window
root = tk.Tk()
root.title("URL Loader")

url_label = tk.Label(root, text="Enter URL:")
url_label.pack()
url_entry = tk.Entry(root, width=50)
url_entry.pack(pady=5)

# Create button to submit URL
submit_button = tk.Button(root, text="Load URL", command=load_url)
submit_button.pack(pady=5)

# Initialize Selenium WebDriver
driver = webdriver.Chrome()

# Read Excel data
excel_data_df = pd.read_excel('Prototype - Copy - Copy.xlsx')

# Run the Tkinter event loop
root.mainloop()

# Quit the WebDriver when the GUI is closed
driver.quit()
